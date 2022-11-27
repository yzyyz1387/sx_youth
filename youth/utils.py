# python3
# -*- coding: utf-8 -*-
# @Time    : 2022/8/14 13:17
# @Author  : yzyyz
# @Email   :  youzyyz1384@qq.com
# @File    : utils.py
# @Software: PyCharm
import json
import logging
import os
import sys
import signal
from openpyxl import load_workbook
from functools import wraps
from os.path import dirname
from typing import Optional, Union
from playwright.async_api import async_playwright, Browser, BrowserContext, Page

import aiofiles
import yaml
from nonebot import logger, get_driver
from jinja2 import Environment, FileSystemLoader
from nonebot.adapters.onebot.v11 import Event, GroupMessageEvent, MessageSegment, PrivateMessageEvent
from playwright.async_api import async_playwright, Browser

from .path import *


def log():
    """
    日志
    :return:
    """
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s-%(levelname)s: %(message)s')
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.formatter = formatter
    console_handler.setLevel(logging.INFO)
    logfile = 'nwafu.log'
    File_handler = logging.FileHandler(logfile, mode='a', encoding='utf-8')
    File_handler.setFormatter(formatter)
    File_handler.setLevel(logging.DEBUG)
    if not logger.handlers:
        logger.addHandler(File_handler)
        logger.addHandler(console_handler)
        print(logger.handlers)
    return logger


async def browser_init(**kwargs) -> Optional[Browser]:
    global _browser
    try:
        browser = await async_playwright().start()
        _browser = await browser.chromium.launch(**kwargs)
        return _browser
    except NotImplementedError:
        logger.warning("windows环境下 初始化playwright失败，相关功能将被限制....")
    except Exception as e:
        logger.warning(f"启动chromium发生错误 {type(e)}：{e}")
        try:
            if _browser:
                await _browser.close()
        except NameError:
            logger.info("正在安装浏览器...")
            os.system("playwright install chromium")
    return None


async def async_w(file, content) -> None:
    """
    异步写入文件
    :param file: 文件
    :param content: 内容
    :return:
    """
    async with aiofiles.open(file, 'w', encoding='utf-8') as f:
        await f.write(content)
        await f.close()


async def async_r(file):
    """
    异步读取文件
    :param file: 文件
    :return: 内容
    """
    async with aiofiles.open(file, 'r', encoding='utf-8') as f:
        content = await f.read()
        await f.close()
        return content


async def init(**kwargs) -> Optional[Browser]:
    global _browser
    try:
        browser = await async_playwright().start()
        _browser = await browser.chromium.launch(**kwargs)
        return _browser
    except NotImplementedError:
        logger.warning("win环境下 初始化playwright失败，相关功能将被限制....")
    except Exception as e:
        logger.warning(f"启动chromium发生错误 {type(e)}：{e}")
        try:
            if _browser:
                await _browser.close()
        except NameError:
            logger.info("正在安装浏览器...")
            os.system("playwright install chromium")
    return None


async def cue_user(event: Event, qq: Optional[int]) -> MessageSegment:
    """
    根据不同事件返回不同的用户提醒方式
    :param event:
    :param qq:
    :return:
    """
    if isinstance(event, GroupMessageEvent):
        return MessageSegment.at(qq)
    elif isinstance(event, PrivateMessageEvent):
        return MessageSegment.text("")


async def plugin_init():
    """
    初始化插件
    :return:
    """
    Path.mkdir(LOCAL) if not Path.exists(LOCAL) else ...
    Path.mkdir(IMG_OUTPUT_PATH) if not Path.exists(IMG_OUTPUT_PATH) else ...
    Path.mkdir(OUTPUT_PATH) if not Path.exists(OUTPUT_PATH) else ...
    Path.mkdir(VERIFY_PATH) if not Path.exists(VERIFY_PATH) else ...
    if LOCK_PATH.exists():
        LOCK_PATH.unlink()
    for i in list(LOCAL.iterdir()):
        if i.suffix in ['.xlsx', '.xls']:
            try:
                workbook_ = load_workbook(i)
                sheet_names = workbook_.sheetnames
                sheet = workbook_[sheet_names[0]]
                youth_qq_info = {}
                for row in sheet.rows:
                    youth_qq_info[row[0].value] = row[1].value
                workbook_.close()
                await async_w(YOUTH_QQ_PATH, json.dumps(youth_qq_info, ensure_ascii=False))
                break
            except:
                logger.info("读取储存有qq信息的本地excel文件失败")
                if YOUTH_QQ_PATH.exists():
                    YOUTH_QQ_PATH.unlink()
    logger.success("陕西共青团插件初始化检测完成")


# async def render_and_shoot(temp_path: str, data, html_write: Union[Path, str], img_output: Union[Path, str]) -> None:
#     env = Environment(loader=FileSystemLoader(str(dirname(__file__))))
#     template = env.get_template(temp_path)
#     html = template.render(data=data)
#     with open(html_write, 'w', encoding='utf-8') as f:
#         f.write(html)
#         f.close()
#     browser = await browser_init()
#     context = await browser.new_context(locale="zh-CN")
#     page = await context.new_page()
#     await page.goto(f"file:///{Path(html_write).resolve()}")
#     await page.locator('.mdui-table').screenshot(path=img_output)


def signal_handler(signum, frame):
    print('signal_handler: caught signal ' + str(signum))
    if signum == signal.SIGINT.value:
        logger.info("操作被用户停止")
        LOCK_PATH.unlink() if LOCK_PATH.exists() else  ...
        sys.exit(1)


signal.signal(signal.SIGINT, signal_handler)


def time_log(func):
    @wraps(func)
    async def wrapper(*args, **kwargs):
        import time
        start_time = time.time()
        data = await func(*args, **kwargs)
        end_time = time.time()
        logger.info(f"运行时间：{end_time - start_time}")
        return data

    return wrapper


async def youth_analyze(data: dict):
    """
    共青团青年大学习分析
    :param data:
    :return:
    """
    logger.info("正在分析大学习数据...")
    data = data['data']['data']
    isStudy = []
    unfinished = []
    for youth in data:
        if youth['isStudy'] == '是':
            isStudy.append(youth['realname'])
        elif youth['isStudy'] == '否':
            unfinished.append(youth['realname'])
    return isStudy, unfinished
